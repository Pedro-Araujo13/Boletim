import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Caminho da pasta com arquivos .txt
caminho_txt = "C:\\Users\\User\\OneDrive\\Documentos\\Programação\\Python\\ConversorPDFtxt\\arquivostxt"

# Caminho do modelo Excel
modelo_excel = "C:\\Users\\User\\OneDrive\\Documentos\\tabela_modelo.xlsx"   # <-- AJUSTE AQUI

# Caminho para salvar o Excel final
saida_excel = "C:\\Users\\User\\OneDrive\\Documentos\\Programação\\Python\\ConversorPDFtxt\\modelo_final.xlsx"  # <-- AJUSTE AQUI

# Nome da turma (definido manualmente aqui)
nome_turma = "TURMA 1°ANO 'A'"

# Lista os arquivos .txt na pasta
txt_files = [f for f in os.listdir(caminho_txt) if f.endswith(".txt")]

# Cria novo workbook final
wb_final = openpyxl.Workbook()
wb_final.remove(wb_final.active)  # Remove aba padrão

# Processa cada .txt
for txt_file in txt_files:
    with open(os.path.join(caminho_txt, txt_file), "r", encoding="utf-8") as f:
        dados = f.read().strip()

    partes = dados.split()
    nome_aluno = " ".join(partes[:-7])  # Assume que as últimas 7 entradas são notas
    notas = list(map(float, partes[-7:]))

    # Carrega o modelo
    modelo_wb = load_workbook(modelo_excel)
    modelo_ws = modelo_wb.active

    # Preenche nome e turma (em células mescladas)
    modelo_ws["D20"] = nome_aluno
    modelo_ws["D21"] = nome_turma

    # Preenche as notas a partir da célula F5, pulando de 2 em 2 linhas
    for i, nota in enumerate(notas):
        linha = 5 + i * 2
        modelo_ws[f"F{linha}"] = nota

    # Cria nova aba no final
    nova_ws = wb_final.create_sheet(title=nome_aluno[:31])  # máximo 31 caracteres

    # Copia valores e estilos com segurança (sem tentar escrever em MergedCell)
    for row in modelo_ws.iter_rows():
        for cell in row:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                nova_cell = nova_ws[cell.coordinate]
                nova_cell.value = cell.value
                nova_cell.font = copy(cell.font)
                nova_cell.border = copy(cell.border)
                nova_cell.fill = copy(cell.fill)
                nova_cell.number_format = copy(cell.number_format)
                nova_cell.protection = copy(cell.protection)
                nova_cell.alignment = copy(cell.alignment)

    # Copia mesclagens
    for merged_range in modelo_ws.merged_cells.ranges:
        nova_ws.merge_cells(str(merged_range))

    # Ajusta automaticamente a largura das colunas
    for col in nova_ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        nova_ws.column_dimensions[col_letter].width = max_length + 2  # margem

# Salva o arquivo final
wb_final.save(saida_excel)
print(f"✅ Arquivo Excel salvo com sucesso em:\n{saida_excel}")
